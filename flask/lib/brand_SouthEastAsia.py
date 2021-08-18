import pandas as pd
from path import UPLOAD_SEA_FORM
from openpyxl import load_workbook 


# def excel2df(file_path, sheet_name, a)
def excel2df(file_path):
    try:
        df = pd.read_excel(file_path)
        df = df.fillna("")
        return df
    except:
        print("엑셀파일을 불러오는데 실패했습니다.")
        return 0


def generate_df(brand,order_columns):
    # 빈 데이터프레임 선언
    data = pd.DataFrame()
    product_code = pd.DataFrame()

    for brand_row_num in range(len(brand)):
        # 사이즈 값이 없을 때 = one size, 색상 값이 없을 떄 = one color
        # 색,사이즈 개수가 2개 이상이면, row 추가해 각 컬럼에 색,사이즈 하나씩만 들어갈 수 있도록.
        if brand.iloc[brand_row_num]["색상"] == "":
            list_col = ["ONE COLOR"]
        else:
            list_col = brand.iloc[brand_row_num]["색상"].split('\n')
        # print(list_col)

        if brand.iloc[brand_row_num]["사이즈"] == "":
            list_size = ["ONE SIZE"]
        else:
            list_size = brand.iloc[brand_row_num]["사이즈"].split('\n')
        # print(list_size)

        if brand.iloc[brand_row_num]["무게(g)"] == "":
            list_weight = ""
        else:
            if len(list_size) == 1 :
                list_weight = [brand.iloc[brand_row_num]["무게(g)"]]
                list_weight = list(map(int, list_weight))
                list_weight[0] = list_weight[0]/1000
            else:
                list_weight = [brand.iloc[brand_row_num]["무게(g)"]]
                list_weight = list(map(str, list_weight))
                list_weight = brand.iloc[brand_row_num]["무게(g)"].split('\n') # 이 부분 다시 확인.
                list_weight = list(map(int, list_weight))
                for i in range(len(list_weight)):
                    list_weight[i] = list_weight[i]/1000
        
        # 색상x사이즈 개수 만큼 row 늘려 값 집어넣기
        for numOFlist_col in range(len(list_col)):
            for numOFlist_size in range(len(list_size)):
                val = {'Product Name' : brand.iloc[brand_row_num]["제품명"],
                        'Product Description' : brand.iloc[brand_row_num]["소재"],
                        'Price' : brand.iloc[brand_row_num]["가격"],
                        'Stock' : brand.iloc[brand_row_num]["재고수량"],
                        'Weight' : list_weight[numOFlist_size],
                        # 'Variation Integration No.' : brand.iloc[brand_row_num]["상품코드"],
                        'Variation Name1' : "COLOR",
                        'Option for Variation 1' : list_col[numOFlist_col],
                        'Variation Name2' : "SIZE",
                        'Option for Variation 2' : list_size[numOFlist_size]}

# 'Category','Maximum Purchase Quantity','Maximum Purchase Quantity - Start Date'
# ,'Maximum Purchase Quantity - Time Period (in Days)','Maximum Purchase Quantity - End Date','Parent SKU'
# ,'Image per Variation'
# ,'SKU','Cover image','Item image 1','Item image 2','Item image 3','Item image 4','Item image 5','Item image 6','Item image 7','Item image 8'
# ,'Length','Width','Height','Standard Express - Korea','Pre-order DTS'

                data = data.append(val, ignore_index=True)
        
        # print(data)

        # data["이미지"] = brand["상품코드"].apply(lambda x : [name for name in jpg_list if x in name])

    # 참조정보가 없는 컬럼 빈칸 처리
    for column in order_columns:
        if (column in data.columns):
            continue
        else:
            data[column] = ""

    data = data[order_columns]
    return data

def df2excel(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['Template']

    # 각 column의 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 6
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt).value = val
            row_cnt += 1
        col_cnt += 1

    wb.save(new_path)

def brand2SEA(file_path,upload_path):
    order_columns = ['Category','Product Name','Product Description','Maximum Purchase Quantity','Maximum Purchase Quantity - Start Date'
                    ,'Maximum Purchase Quantity - Time Period (in Days)','Maximum Purchase Quantity - End Date','Parent SKU','Variation Integration No.'
                    ,'Variation Name1','Option for Variation 1','Image per Variation','Variation Name2','Option for Variation 2','Price','Stock'
                    ,'SKU','Cover image','Item image 1','Item image 2','Item image 3','Item image 4','Item image 5','Item image 6','Item image 7','Item image 8'
                    ,'Weight','Length','Width','Height','Standard Express - Korea','Pre-order DTS']
                    
    brand_df = excel2df(file_path)
    # 정보 나타내는 행들 제거
    brand_df.drop([brand_df.index[0],brand_df.index[1]],inplace=True)
    # print(brand_df)
    brand_df = brand_df.reset_index()
    sea_df = generate_df(brand_df,order_columns)
    df2excel(sea_df,UPLOAD_SEA_FORM,upload_path)