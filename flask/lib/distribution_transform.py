import pandas as pd
from openpyxl import load_workbook
from path import DISTRIBUTION_EXCEL_FORM

# 주문정보->물류파일 변환
# 파일 불러오기


def order_info(file_path, sheet_name=None):
    if sheet_name:
        try:
            df = pd.read_excel(file_path, sheet_name)
            return df
        except:
            print("엑셀파일을 불러오는데 실패했습니다.")
            return 0

    try:
        df = pd.read_csv(file_path)
        return 0
    except:
        print("csv파일을 불러오는데 실패했습니다.")
        return 0


# 주문정보에 있는 data, distribution으로 가져오기
def distribution_df(order_info_df, form_data, distribution_columns):
    distribution_dict = dict()
    for column in order_info_df.columns:
        data = order_info_df[column]
        if (column == "order No."):
            distribution_dict["Shipment Reference No."] = data
            distribution_dict["Shipment Reference No.2"] = data
        elif (column == "customer name"):
            distribution_dict["Receiver Contact Name"] = data
        elif (column == "customer phone"):
            distribution_dict["Receiver Tel"] = data
        elif (column == "detailed address"):
            distribution_dict["Receiver Address"] = data
        elif (column == "city"):
            distribution_dict["Receiver City"] = data
        elif (column == "province"):
            distribution_dict["Receiver Province"] = data
        elif (column == "certificates NO."):
            distribution_dict["Receiver Credentials No"] = data
        elif (column == "vendor remark"):
            distribution_dict["Receiver Postal Code"] = data

    for key in form_data:
        if form_data[key] == "":
            continue
        else:
            distribution_dict[key] = [form_data[key]] * len(order_info_df)

    new_df = pd.DataFrame(distribution_dict)


# 얻지 못하는 값들은 빈 값 처리
    for column in distribution_columns:
        if (column in new_df.columns):
            continue
        else:
            new_df[column] = ""

    # 물류파일 컬럼순으로 정렬
    New_distribution_df = new_df[distribution_columns]

    return New_distribution_df


def append_value(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['물류파일']

    # 각 column 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 4
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt).value = val
            row_cnt += 1
        col_cnt += 1
    # save
    wb.save(new_path)


def distribution_from_orderinfo(file, form_data, upload_path):
    # sheet_name = {'발주':'발주파일', '물류':'물류파일', '상품':'commodity', '주문[영문]':'secoo주문영문', '주문[중문]': 'secoo주문중문', '마스터': '마스타파일'}
    # need_columns = ['order No.', 'customer name', 'customer phone', 'detailed address', 'city', 'province', 'certificates NO.', 'vedor remark']
    distribution_columns = ['Shipment Reference No.', 'Shipment Reference No.2', 'Customer Account No.', 'Company Name', 'Contact Name', 'Tel',
                            'Address', 'City', 'Province', 'Country/Region', 'Email', 'Postal Code', 'Receiver Customer Account No.', "Receiver's Tax ID No.",
                            'Receiver Company Name', 'Receiver Contact Name', 'Receiver  Chinese Name', 'Receiver Tel', 'Receiver Address', 'Receiver City', 'Receiver Province',
                            'Receiver Email', 'Receiver Country/Region', 'Receiver Credentials Type', 'Receiver Credentials No', 'Receiver Postal Code', 'Currency', 'Shipment Type',
                            'Add Service1', 'Add Service Account1', 'Add Service Amount1', 'Add Service2', 'Add Service Account2', 'Add Service Amount2', 'Total Package', 'Self Pickup',
                            'Payment Method', 'Account No.', 'Third Party District Code', 'Tax payment', '(PCS) 1', '(L) 1', '(W) 1', '(H) 1', 'Term of Trade', 'Reason for Sending 1', 'Reason for Sending 2',
                            'Remarks', 'AWB Remarks', 'Tax Account', 'Add Service3', 'VAT', 'EORI', 'Traceability label', 'PO Number', "Shipper's Tax ID No.", 'Appointment Time', 'Personal baggage', 'Importer Company Name',
                            'Importer Tel No.', 'Importer Address', 'Agent Waybill', 'Customs Clearance', 'Shipper County', 'Receiver County', 'Shipper Credentials Type', 'Shipper Credentials No.', 'Notify to pick up',
                            'Formal Customs Declaration For Import']

    order_info_df = order_info(file, sheet_name=form_data["sheet"])

    New_distribution_df = distribution_df(
        order_info_df, form_data, distribution_columns)
    append_value(New_distribution_df, DISTRIBUTION_EXCEL_FORM, upload_path)
