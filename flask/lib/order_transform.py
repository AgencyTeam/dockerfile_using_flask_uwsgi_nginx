from openpyxl import load_workbook
import pandas as pd
from path import ORDER_EXCEL_FORM


# excel file -> df
def excel2df(file_path, sheet_name):
    try:
        df = pd.read_excel(file_path, sheet_name)
        return df
    except:
        print("엑셀파일을 불러오는데 실패했습니다.")
        return 0


# 주문정보 df 정제
def df_transform(order_df, master_df, need_columns):
    for column in order_df:
        if (column in need_columns):
            continue
        else:
            order_df = order_df.drop(column, axis=1)

    # 판매가, 배송비 column 생성
    customer_price = []
    delivery_price = []
    for product_code in order_df["product model"]:
        is_code = master_df["자체 상품코드"] == product_code
        value1 = list(master_df[is_code]["판매가"])
        value2 = list(master_df[is_code]["배송비"])
        customer_price.append(value1[0]) if value1 != [
        ] else customer_price.append(0)
        delivery_price.append(value2[0]) if value2 != [
        ] else delivery_price.append(0)
    order_df["소비자가"] = customer_price
    order_df["배송비"] = delivery_price

    return order_df


# 정제된 발주파일 df 생성
def generate_df(transformed_order_df, form_data, order_columns):
    order_dict = dict()

    for column in transformed_order_df.columns:
        data = transformed_order_df[column]
        if (column == "product model"):
            order_dict["상품코드"] = data
        elif (column == "vendor product No."):
            order_dict["사이즈코드"] = list(map(lambda x: x[-3:], data))
        elif (column == "product quantity"):
            order_dict["주문수량"] = data
        elif (column == "order No."):
            order_dict["외부몰주문번호"] = data
            order_dict["주문 메모"] = data
        elif (column == "create time"):
            tmp = []
            for date in data:
                ymd = date.split(' ')
                tmp.append(ymd[0])
            order_dict["결제일시"] = tmp
        elif (column == "settle price"):
            data = list(map(int, data))
            order_dict["실판매가"] = list(map(lambda x: x/67*100, data))
        elif (column == "소비자가"):
            order_dict["소비자가"] = data
        elif (column == "배송비"):
            order_dict["배송비"] = data
    order_dict["실판매가-배송비"] = [a - b for a,
                              b in zip(order_dict["실판매가"], order_dict["배송비"])]
    order_dict["발주가"] = [b if a > b else a for a, b in zip(
        order_dict["실판매가-배송비"], order_dict["소비자가"])]
    order_dict["발주가(최종)"] = list(map(lambda x: x - x %
                                     10, map(int, order_dict["발주가"])))
    order_dict["총주문금액"] = order_dict["발주가(최종)"]

    # form 에서 온 데이터들 추가
    for key in form_data:
        if form_data[key] == "":
            continue
        else:
            if key == "환율":
                order_dict[key] = [int(form_data[key])] * \
                    len(transformed_order_df)
            else:
                order_dict[key] = [form_data[key]] * len(transformed_order_df)

    # 만든 dict -> df로 변환
    new_df = pd.DataFrame(order_dict)

    # 얻지 못하는 값들은 빈 값 처리
    for column in order_columns:
        if (column in new_df.columns):
            continue
        else:
            new_df[column] = ""

    # 발주파일 column 순서대로 정렬
    new_df = new_df[order_columns]

    return new_df


def df2excel(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['발주파일']

    # 각 column의 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 3
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt).value = val
            row_cnt += 1
        col_cnt += 1

    # save
    wb.save(new_path)


def make_excel(file, form_data, upload_path):
    need_columns = ['product model', 'vendor product No.',
                    'product quantity', 'order No.', 'create time', 'settle price']
    order_columns = ['상품코드', '사이즈코드', '주문수량', '외부몰주문번호', '총주문금액', '결제일시', '상점ID', '주문자 이름', '주문자 전화번호',
                     '주문자 휴대폰', '주문자 이메일', '주문자 우편번호', '주문자 고정주소', '주문자 상세주소', '수령자 이름', '수령자 전화번호', '수령자 휴대폰',
                     '수령자 이메일', '수령자 우편번호', '수령자 고정주소', '수령자 상세주소', '주문 메모', '업체 코드', '외부몰 부주문 코드', '환율',
                     'LF CJ운송장', 'SF EXPRESS 운송장', '소비자가', '실판매가', '배송비', '실판매가-배송비', '발주가', '발주가(최종)']

    order_df = excel2df(file, sheet_name=form_data["주문시트"])
    master_df = excel2df(file, sheet_name=form_data["마스타시트"])

    transformed_order_df = df_transform(order_df, master_df, need_columns)

    order = generate_df(transformed_order_df, form_data, order_columns)

    df2excel(order, ORDER_EXCEL_FORM, upload_path)
