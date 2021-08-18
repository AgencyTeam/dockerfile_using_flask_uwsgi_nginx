import pandas as pd
from path import UPLOAD_DOMESTIC_FORM
from openpyxl import load_workbook 


def excel2df(file_path):
    try:
        df = pd.read_excel(file_path)
        return df
    except:
        print("엑셀파일을 불러오는데 실패했습니다.")
        return 0

def find_list(jpg_list,name):
    return ','.join([i for i in jpg_list if i.startswith(name) ])

def generate_df(brand,order_columns):
    # 빈 데이터프레임 선언
    data = pd.DataFrame()

    # 통합데이터포맷 값 직접사용
    data["브랜드"] = brand["브랜드명"].str.strip()
    data["자체 상품코드"] = brand["상품코드"].str[0:50]
    data["상품명"] = brand["제품명"].str[0:100]
    data["필수옵션명"] = "색상"
    data["필수옵션값"] = brand["색상"]
    data["필수옵션가"] = brand["가격"]
    data["선택옵션명"] = "사이즈"
    data["선택옵션값"] = brand["사이즈"]
    data["선택옵션가"] = brand["가격"]
    data["판매가"] = brand["가격"]
    data["무게"] = brand["무게(g)"]
    data["제조사"] = brand["제조사"].str.strip()
    data["원산지"] = brand["원산지"]
    data["요약설명"] = brand["상품설명"]
    data["상품 상세정보"] = brand["상세정보(html)"]
    data["재고수량"] = brand["재고수량"]

    data["판매상태"] = brand["재고수량"].apply(lambda x: "판매중" if x > 0 else "품절")
    # data["대표 이미지 파일명"] = brand["상품코드"].apply(lambda x : find_list(jpg_list,x))

    # 미입력시 기본값 적용되는 컬럼들.
    data["상품상태"] = "신상품"
    data["재고사용"] = "Y"
    data["세금"] = "과세상품"
    data["미성년자 구매"] = "Y"
    data["묶음배송 가능"] = "Y"
    data["별도설치비"] = "N"
    data["지역별 배송비 사용 여부"] = "A"
    data["옵션형태"] = "조합형"
    data["재고소진후주문가능여부"] = "N"
    # data["네이버/다음 쇼핑 노출용 상품명"] = brand["상품명"].str[0:100]
    data["주문제작상품"] = "N"
    data["병행수입"] = "N"
    data["해외구매대행"] = "N"
    data["판매방식"] = "소매"
    data["다음 쇼핑하우 노출 설정"] = "N"
    data["네이버 쇼핑 노출 설정"] = "N"
    data["네이버 페이 구매가능 설정"] = "N"
    data["Facebook 다이내믹 광고 설정"] = "N"


    # 참조정보가 없는 컬럼 빈칸 처리
    for column in order_columns:
        if (column in data.columns):
            continue
        else:
            data[column] = ""

    data = data[order_columns]
    return data

def df2excel(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['ver.1.1']

    # 각 column의 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 4
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt).value = val
            row_cnt += 1
        col_cnt += 1

    wb.save(new_path)

def brand2domestic(file_path,upload_path):
    order_columns = ['상품명','자체 상품코드','카테고리ID','요약설명','판매상태','상품상태','판매가','무게','정가'
                        ,'재고사용','재고수량','SKU(재고번호)','대표 이미지 파일명','상품 상세정보','세금'
                        ,'미성년자 구매','개인통관고유부호','원산지','제조사','브랜드','배송방법','택배 가능 여부'
                        ,'퀵서비스 가능 여부','방문 수령 가능 여부','배송비 결제 방법','배송비 유형','기본배송비'
                        ,'조건부무료-상품판매가합계','무게별 차등 배송비(기본가격)','무게별 차등 배송비(무게)'
                        ,'무게별 차등 배송비(가격)','수량별 차등 배송비(타입)','수량별 차등 배송비(기본 가격)'
                        ,'수량별 차등 배송비(수량)','수량별 차등 배송비(가격)','수량별 차등 배송비(반복수량)'
                        ,'수량별 차등 배송비(반복가격)','구매금액별 차등 배송비(기본 가격)','구매금액 차등 배송비(구매금액)'
                        ,'구매금액별 차등 배송비(가격)','묶음배송 가능','별도설치비','지역별 배송비 사용 여부','지역별 배송비 유형'
                        ,'간편설정(제주도 추가 배송비)','간편설정(도서산간 추가 배송비)','우편번호 등록(시작구간)'
                        ,'우편번호 등록(종료구간)','우편번호 등록(추가배송비)','상품구매시 적립금 지급 값','상품구매시 적립금 지급 단위'
                        ,'할인적용대상','할인금액','할인금액 단위','옵션형태','필수 옵션명','필수 옵션값','필수 옵션가'
                        ,'옵션 재고수량','선택 옵션명','선택 옵션값','선택 옵션가','선택 옵션 재고수량','사용자 입력형 옵션'
                        ,'0원 선택옵션 최대 구매수량','재고소진후주문가능여부','네이버/다음 쇼핑 노출용 상품명','네이버 쇼핑 이벤트 문구'
                        ,'네이버 쇼핑 카테고리 ID','최소 구매수량','1회 구매시 최대 수량','1인 최대 구매수량'
                        ,'주문제작상품','병행수입','해외구매대행','판매방식','다음 쇼핑하우 노출 설정','네이버 쇼핑 노출 설정'
                        ,'네이버 페이 구매가능 설정','Facebook 다이내믹 광고 설정']
    
    brand_df = excel2df(file_path)
    # 이미지컬럼 제거
    del brand_df["이미지"]
    # 정보 나타내는 행들 제거
    brand_df.drop([brand_df.index[0],brand_df.index[1]],inplace=True)

    domestic_df = generate_df(brand_df, order_columns)
    # print(domestic_df.head(5))
    df2excel(domestic_df,UPLOAD_DOMESTIC_FORM,upload_path)
