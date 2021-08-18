import pandas as pd
from pandas.core import base
from path import UPLOAD_CHINA_FORM
from openpyxl import load_workbook

def excel2df(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name='위챗미니프로그램')
        df = df.fillna("")
        return df
    except:
        print("엑셀파일을 불러오는데 실패했습니다.")
        return 0

def generate_df(brand,order_columns):
    # 빈 데이터프레임 선언
    data = pd.DataFrame()

    for brand_row_num in range(len(brand)):
        # 사이즈 값이 없을 때 = one size, 색상 값이 없을 떄 = one color
        # 색,사이즈 개수가 2개 이상이면, row 추가해 각 컬럼에 색,사이즈 하나씩만 들어갈 수 있도록.
        if brand.iloc[brand_row_num]["색상"] == "":
            list_col = ["ONE COLOR"]
        else:
            list_col = brand.iloc[brand_row_num]["색상"].split('\n')
        print(list_col)

        if brand.iloc[brand_row_num]["사이즈"] == "":
            list_size = ["ONE SIZE"]
        else:
            list_size = brand.iloc[brand_row_num]["사이즈"].split('\n')
        print(list_size)

        #number of list
        for nol_col in range(len(list_col)):
            for nol_size in range(len(list_size)):
                val = {'상품명' : brand.iloc[brand_row_num]['제품명'],
                        '구분(품번)' : brand.iloc[brand_row_num]['상품코드'],
                        '정상공급가' : brand.iloc[brand_row_num]['가격'],
                        '재고수량' : brand.iloc[brand_row_num]['재고수량'],
                        '원산지(제조국)(영문)' : brand.iloc[brand_row_num]['원산지'] + "/Korea",
                        '소재' : brand.iloc[brand_row_num]['소재'], 
                        '세탁방법' : brand.iloc[brand_row_num]['세탁방법'],
                        '제품 설명' : brand.iloc[brand_row_num]['상품설명'],
                        '색상(영문)' : list_col[nol_col],
                        '사이즈' : list_size[nol_size],
                        '상의-사이즈' : list_size[nol_size],
                        '하의-사이즈' : list_size[nol_size]}
                
                data = data.append(val, ignore_index=True)
    # 중국 서버용 형식에 맞게 변환 / 정확해보인것들 일부만 해봄.
    #data["구분(품번)"] = brand["상품코드"]
    # 브랜드에서 한글로 제품명 제공한경우, 영문으로 바꿔야함
    #data["상품명"] = brand["제품명"]

    #data["정상공급가"] = brand["가격"]
    #data["색상(영문)"] = brand["색상"]
    #data["사이즈"] = brand["사이즈"]
    #data["재고수량"] = brand["재고수량"]
    #data["원산지(제조국)(영문)"] = brand["원산지"] + "/Korea"
    #data["세탁방법"] = brand["세탁방법"]
    #data["소재"] = brand["소재"]
    #data['스타일+사이즈'] = brand["상품코드"].str[0:8] + data['사이즈']
    #data['제품 설명'] = brand['상품설명']

    

    # 참조정보가 없는 컬럼 빈칸 처리
    for column in order_columns:
        if (column in data.columns):
            continue
        else:
            data[column] = ""

    data = data[order_columns]
    return data

def df2excel(df, form_path, new_path):
    wb = load_workbook(form_path)
    ws = wb['IMFORM']

    # 각 column의 값 추가
    col_cnt = 1
    for col in df.columns:
        row_cnt = 8
        for val in df[col]:
            ws.cell(row=row_cnt, column=col_cnt+1).value = val
            row_cnt += 1
        col_cnt += 1

    wb.save(new_path)

def brand2china(file_path,upload_path):
    order_columns = ['구분(품번)','상품명','정상공급가','할인율','할인공급가','색상(영문)','사이즈','사이즈(물산)','재고수량'
                    ,'원산지(제조국)(영문)','카테고리(대분류)','카테고리(중분류)','카테고리(소분류)','스타일+사이즈'
                    ,'상의-사이즈','상의-어깨너비','상의-가슴너비','상의-소매길이','상의-총장(앞)'
                    ,'하의-사이즈','하의-총장(아웃심)','하의-허리','하의-엉덩이','하의-허벅지','하의-밑위','하의-밑단'
                    ,'세탁방법','품목 및 모델명','소재','제품 설명']

    brand_df = excel2df(file_path)
    del brand_df["이미지"]

    brand_df.drop([brand_df.index[0], brand_df.index[1]],inplace=True)
    brand_df = brand_df.reset_index()
    china_df = generate_df(brand_df, order_columns)

    df2excel(china_df,UPLOAD_CHINA_FORM,upload_path)
